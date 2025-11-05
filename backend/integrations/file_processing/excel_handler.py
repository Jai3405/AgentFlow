"""
Excel file processing handler
Read, write, and manipulate Excel files (.xlsx, .xls)
"""

from typing import List, Dict, Any, Optional, Union
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel support disabled.")

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class ExcelHandler:
    """Handler for Excel file operations"""

    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            print("Warning: Install openpyxl for Excel support: pip install openpyxl")

    async def read_excel(
        self,
        file_path: str,
        sheet_name: str = None,
        has_header: bool = True,
        skip_rows: int = 0
    ) -> Union[List[Dict[str, Any]], Dict[str, List[Dict[str, Any]]]]:
        """
        Read Excel file

        Args:
            file_path: Path to Excel file
            sheet_name: Specific sheet name (if None, reads all sheets)
            has_header: Whether first row is header
            skip_rows: Number of rows to skip from top

        Returns:
            List of row dicts (single sheet) or dict of sheet_name -> rows (all sheets)
        """
        if not OPENPYXL_AVAILABLE:
            raise Exception("openpyxl not installed. Install with: pip install openpyxl")

        try:
            workbook = load_workbook(file_path, data_only=True)

            if sheet_name:
                # Read specific sheet
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"Sheet '{sheet_name}' not found")
                sheet = workbook[sheet_name]
                return await self._read_sheet(sheet, has_header, skip_rows)
            else:
                # Read all sheets
                result = {}
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    result[sheet_name] = await self._read_sheet(sheet, has_header, skip_rows)
                return result

        except Exception as e:
            raise Exception(f"Failed to read Excel file: {str(e)}")

    async def _read_sheet(
        self,
        sheet,
        has_header: bool,
        skip_rows: int
    ) -> List[Dict[str, Any]]:
        """Read data from a worksheet"""
        rows = []
        headers = None

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            # Skip initial rows
            if i < skip_rows:
                continue

            # Get headers
            if i == skip_rows and has_header:
                headers = [str(cell) if cell is not None else f'Column_{j}' for j, cell in enumerate(row)]
                continue

            # Process data rows
            if headers:
                row_dict = {}
                for j, cell in enumerate(row):
                    if j < len(headers):
                        row_dict[headers[j]] = cell
                rows.append(row_dict)
            else:
                # No headers, use column indices
                row_dict = {f'Column_{j}': cell for j, cell in enumerate(row)}
                rows.append(row_dict)

        return rows

    async def write_excel(
        self,
        data: Union[List[Dict[str, Any]], Dict[str, List[Dict[str, Any]]]],
        file_path: str,
        sheet_name: str = "Sheet1",
        include_header: bool = True,
        auto_width: bool = True,
        header_style: bool = True
    ) -> bool:
        """
        Write data to Excel file

        Args:
            data: List of dicts (single sheet) or dict of sheet_name -> data (multiple sheets)
            file_path: Output file path
            sheet_name: Sheet name (for single sheet data)
            include_header: Whether to include header row
            auto_width: Auto-adjust column widths
            header_style: Apply styling to header row

        Returns:
            True if successful
        """
        if not OPENPYXL_AVAILABLE:
            raise Exception("openpyxl not installed. Install with: pip install openpyxl")

        try:
            workbook = Workbook()
            # Remove default sheet
            workbook.remove(workbook.active)

            if isinstance(data, dict):
                # Multiple sheets
                for sheet_name, sheet_data in data.items():
                    await self._write_sheet(
                        workbook,
                        sheet_name,
                        sheet_data,
                        include_header,
                        auto_width,
                        header_style
                    )
            else:
                # Single sheet
                await self._write_sheet(
                    workbook,
                    sheet_name,
                    data,
                    include_header,
                    auto_width,
                    header_style
                )

            workbook.save(file_path)
            return True

        except Exception as e:
            raise Exception(f"Failed to write Excel file: {str(e)}")

    async def _write_sheet(
        self,
        workbook: Workbook,
        sheet_name: str,
        data: List[Dict[str, Any]],
        include_header: bool,
        auto_width: bool,
        header_style: bool
    ):
        """Write data to a worksheet"""
        if not data:
            return

        sheet = workbook.create_sheet(title=sheet_name)

        # Get headers
        headers = list(data[0].keys())

        # Write header
        if include_header:
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_idx, value=header)

                # Apply header styling
                if header_style:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        start_row = 2 if include_header else 1
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, header in enumerate(headers, start=1):
                value = row_data.get(header)
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        if auto_width:
            for col_idx, header in enumerate(headers, start=1):
                max_length = len(str(header))
                for row_data in data:
                    value = row_data.get(header)
                    if value:
                        max_length = max(max_length, len(str(value)))
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

    async def read_excel_pandas(
        self,
        file_path: str,
        sheet_name: str = None
    ) -> Union[List[Dict[str, Any]], Dict[str, List[Dict[str, Any]]]]:
        """
        Read Excel using pandas (alternative method)

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name or None for all sheets

        Returns:
            Data as list of dicts
        """
        if not PANDAS_AVAILABLE:
            raise Exception("pandas not installed. Install with: pip install pandas openpyxl")

        try:
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                return df.to_dict('records')
            else:
                # Read all sheets
                excel_file = pd.ExcelFile(file_path)
                result = {}
                for sheet in excel_file.sheet_names:
                    df = pd.read_excel(file_path, sheet_name=sheet)
                    result[sheet] = df.to_dict('records')
                return result

        except Exception as e:
            raise Exception(f"Failed to read Excel with pandas: {str(e)}")

    async def write_excel_pandas(
        self,
        data: Union[List[Dict[str, Any]], Dict[str, List[Dict[str, Any]]]],
        file_path: str,
        sheet_name: str = "Sheet1"
    ) -> bool:
        """
        Write Excel using pandas (alternative method)

        Args:
            data: Data to write
            file_path: Output file path
            sheet_name: Sheet name (for single sheet)

        Returns:
            True if successful
        """
        if not PANDAS_AVAILABLE:
            raise Exception("pandas not installed. Install with: pip install pandas openpyxl")

        try:
            if isinstance(data, dict):
                # Multiple sheets
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    for sheet, sheet_data in data.items():
                        df = pd.DataFrame(sheet_data)
                        df.to_excel(writer, sheet_name=sheet, index=False)
            else:
                # Single sheet
                df = pd.DataFrame(data)
                df.to_excel(file_path, sheet_name=sheet_name, index=False)

            return True

        except Exception as e:
            raise Exception(f"Failed to write Excel with pandas: {str(e)}")

    async def get_sheet_names(self, file_path: str) -> List[str]:
        """
        Get list of sheet names in Excel file

        Args:
            file_path: Path to Excel file

        Returns:
            List of sheet names
        """
        if not OPENPYXL_AVAILABLE:
            raise Exception("openpyxl not installed")

        try:
            workbook = load_workbook(file_path, read_only=True)
            return workbook.sheetnames
        except Exception as e:
            raise Exception(f"Failed to read sheet names: {str(e)}")

    async def transform_excel(
        self,
        file_path: str,
        transformations: List[Dict[str, Any]],
        output_path: str,
        sheet_name: str = None
    ) -> bool:
        """
        Apply transformations to Excel data

        Args:
            file_path: Input file path
            transformations: List of transformation operations
            output_path: Output file path
            sheet_name: Sheet to transform (None for all)

        Returns:
            True if successful
        """
        # Read Excel
        data = await self.read_excel(file_path, sheet_name)

        # Apply transformations to each sheet
        if isinstance(data, dict):
            transformed = {}
            for sheet, sheet_data in data.items():
                transformed[sheet] = await self._apply_transformations(sheet_data, transformations)
        else:
            transformed = await self._apply_transformations(data, transformations)

        # Write transformed data
        await self.write_excel(transformed, output_path)
        return True

    async def _apply_transformations(
        self,
        data: List[Dict[str, Any]],
        transformations: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """Apply transformations to data"""
        result = data

        for transform in transformations:
            operation = transform.get('operation')

            if operation == 'filter':
                condition = transform.get('condition')
                result = [row for row in result if self._evaluate_condition(row, condition)]

            elif operation == 'select':
                columns = transform.get('columns', [])
                result = [{col: row.get(col) for col in columns} for row in result]

            elif operation == 'rename':
                mapping = transform.get('mapping', {})
                result = [{mapping.get(k, k): v for k, v in row.items()} for row in result]

            elif operation == 'sort':
                sort_by = transform.get('column')
                reverse = transform.get('descending', False)
                result = sorted(result, key=lambda x: x.get(sort_by, ''), reverse=reverse)

        return result

    def _evaluate_condition(self, row: Dict[str, Any], condition: Dict[str, Any]) -> bool:
        """Evaluate filter condition"""
        column = condition.get('column')
        operator = condition.get('operator')
        value = condition.get('value')

        row_value = row.get(column)

        if operator == 'equals':
            return row_value == value
        elif operator == 'not_equals':
            return row_value != value
        elif operator == 'contains':
            return value in str(row_value)
        elif operator == 'greater_than':
            try:
                return float(row_value) > float(value)
            except:
                return False
        elif operator == 'less_than':
            try:
                return float(row_value) < float(value)
            except:
                return False
        else:
            return True

    async def merge_excel_files(
        self,
        file_paths: List[str],
        output_path: str,
        merge_type: str = 'sheets'
    ) -> bool:
        """
        Merge multiple Excel files

        Args:
            file_paths: List of Excel file paths
            output_path: Output file path
            merge_type: 'sheets' (combine as separate sheets) or 'data' (combine data)

        Returns:
            True if successful
        """
        if merge_type == 'sheets':
            # Combine as separate sheets
            workbook = Workbook()
            workbook.remove(workbook.active)

            for file_path in file_paths:
                data = await self.read_excel(file_path)
                if isinstance(data, dict):
                    for sheet_name, sheet_data in data.items():
                        await self._write_sheet(workbook, sheet_name, sheet_data, True, True, True)
                else:
                    sheet_name = Path(file_path).stem
                    await self._write_sheet(workbook, sheet_name, data, True, True, True)

            workbook.save(output_path)
            return True

        elif merge_type == 'data':
            # Combine all data into single sheet
            combined_data = []
            for file_path in file_paths:
                data = await self.read_excel(file_path)
                if isinstance(data, dict):
                    for sheet_data in data.values():
                        combined_data.extend(sheet_data)
                else:
                    combined_data.extend(data)

            await self.write_excel(combined_data, output_path, sheet_name="Combined")
            return True

        else:
            raise ValueError(f"Unknown merge_type: {merge_type}")
